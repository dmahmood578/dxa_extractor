"""
dxa_to_excel.py
===============
Reads pre-extracted OCR text from extracted_text/Patient_N/ and writes a
clean, colour-coded Excel spreadsheet to dxa_data.xlsx.

ALL data stays local — no network calls.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── paths ─────────────────────────────────────────────────────────────────────

_SCRIPT_DIR = Path(__file__).parent
TEXT_DIR    = _SCRIPT_DIR / "extracted_text"
OUTPUT_XLS  = _SCRIPT_DIR / "dxa_data.xlsx"

# ── constants ─────────────────────────────────────────────────────────────────

BMD_MIN, BMD_MAX     = 0.3, 2.5
SCORE_MIN, SCORE_MAX = -6.0, 6.0

T_OSTEOPOROSIS = -2.5   # ≤ this → red
T_OSTEOPENIA   = -1.0   # ≤ this → yellow


def _build_ocr_sign_map() -> dict[str, float]:
    """
    Build the OCR→float correction table programmatically.
    GE Lunar tables use 2-char tokens where a letter prefix encodes the integer
    part of a negative score and a digit encodes the tenth:
      O/o = −0.X,  a/A = −1.X,  B/b = −3.X
    """
    m: dict[str, float] = {}
    for d in range(1, 10):
        for prefix in ("O", "o"):
            m[f"{prefix}{d}"] = -d / 10.0
    for d in range(1, 10):
        for prefix in ("a", "A"):
            m[f"{prefix}{d}"] = -(1.0 + d / 10.0)
    m.update({
        "OT": -0.1, "ot": -0.1, "ut": -0.1, "UT": -0.1,
        "Os": -0.5, "os": -0.5,
        "al": -1.2, "AL": -1.2,
        "as": -1.5, "AS": -1.5,
        "AT": -2.1, "aT": -2.1,
        "af": -2.5, "AF": -2.5,
        "Bl": -3.1, "BI": -3.1,
    })
    return m

_OCR_SIGN_MAP: dict[str, float] = _build_ocr_sign_map()

# ── pre-compiled regexes ──────────────────────────────────────────────────────

_RE_AGE_LABEL   = re.compile(r"\bage\s*:\s*([0-9.]+)", re.IGNORECASE)
_RE_AGE_YEARS   = re.compile(r"(\d{2}\.?\d)\s*years", re.IGNORECASE)
_RE_DOB         = re.compile(r"(?:birth\s*date|dob)\s*[:\s]+([0-9/A-Za-z]+)", re.IGNORECASE)
_RE_DOB_DIGITS  = re.compile(r"(\d{1,2})[\/-](\d{1,2})[\/-](\d{2,4})")
_RE_NAME        = re.compile(r"(?:patient|name)\s*:\s*['\"''""\s]*([A-Za-z][A-Za-z\s,.'\-]{3,40})", re.IGNORECASE)
_RE_SEX_LABEL   = re.compile(r"(?:sex|sexe)\s*[:/]\s*(female|male|f|m)", re.IGNORECASE)
_RE_SEX_WORD    = re.compile(r"\b(Female|Male)\b", re.IGNORECASE)
_RE_HEIGHT_IN   = re.compile(r"(?:height|heigl)\s*[/:\s]+([0-9a-zA-Z.]{3,7})\s*in", re.IGNORECASE)
_RE_HEIGHT_UNIT = re.compile(r"height.*?\(in\)\s*([0-9.]+)", re.IGNORECASE)
_RE_WEIGHT      = re.compile(r"weight\s*[/:\s]+([0-9.]+)\s*(?:lb|kg|Ib|bs)", re.IGNORECASE)
_RE_SCAN_DATE   = re.compile(r"(?:measured|scan\s*date)\s*:\s*([0-9A-Za-z/,\s:.-]+?)(?:\s*\(|\s*AM|\s*PM|\n)", re.IGNORECASE)
_RE_DATE_MDY    = re.compile(r"(\d{1,2}/\d{1,2}/(?:20|19)\d{2})")
_RE_PHYSICIAN   = re.compile(r"referring\s*physician\s*:\s*([A-Za-z\s.,'\-]{4,50})", re.IGNORECASE)
_RE_TBS         = re.compile(r"tbs\s*l1[-–]l4\s*:\s*([0-9.]+)", re.IGNORECASE)
_RE_BMD_NEW     = re.compile(r"^[01]\.\d{2,3}")
_RE_BMD_OLD     = re.compile(r"^[01]\d{3}$")
_RE_MERGED_YA_T = re.compile(r"^(\d{2,3})(-\d{1,3})$")
_RE_ANCILLARY   = re.compile(r"ancillary results", re.IGNORECASE)
_RE_WHITESPACE  = re.compile(r"\s+")

# Hologic-specific
_RE_HOLOGIC_BMD_ROW = re.compile(
    r"(?:L1[-–]L4|Neck|Total)\b[^\n]*?(0\.[3-9]\d{2}|1\.\d{3})\s+([-]?\d+\.\d+)(?:\s+([-]?\d+\.\d+))?",
    re.IGNORECASE,
)

_SECTION_STOP_KEYWORDS = frozenset([
    "statistical", "satay", "statay", "stacy", "stascal", "usa", "nhanes",
    "filename", "comments", "tbs", "trend", "frax", "page:", "hologic", "lunar", "©",
])

_SIDE_ALIASES: dict[str, list[str]] = {
    "left":  ["left", "lett", "let", "lef", "lefe"],
    "right": ["right", "righ"],
}

_NAME_STOPS = ["Referring", "Facility", "Birth", "Date", "Patient", "ID", "Phone", "Height"]

# ── data structures ───────────────────────────────────────────────────────────

@dataclass
class BmdResult:
    bmd: Optional[float] = None
    t:   Optional[float] = None
    z:   Optional[float] = None

    def is_valid(self) -> bool:
        return self.bmd is not None and BMD_MIN < self.bmd < BMD_MAX


@dataclass
class FemurResult:
    neck:  BmdResult = field(default_factory=BmdResult)
    total: BmdResult = field(default_factory=BmdResult)


# ── text helpers ──────────────────────────────────────────────────────────────

def read_folder(patient_dir: Path) -> dict[str, str]:
    return {
        p.name: p.read_text(encoding="utf-8", errors="replace")
        for p in sorted(patient_dir.glob("*.txt"))
    }


def combined(texts: dict[str, str]) -> str:
    return "\n".join(texts.values())


def _digit_count(s: str) -> int:
    return sum(c.isdigit() for c in s)


# ── OCR score parser ──────────────────────────────────────────────────────────

def parse_score_token(tok: str, expect_negative: bool = False) -> Optional[float]:
    tok = tok.strip("[]|\"'()")
    if not tok or tok in ("N/A", "-", "=", "*", ">", "<"):
        return None
    if tok in _OCR_SIGN_MAP:
        return _OCR_SIGN_MAP[tok]
    # Handle explicit negative sign
    has_explicit_minus = tok.startswith("-")
    cleaned = re.sub(r"[^0-9.\-]", "", tok)
    if not cleaned or cleaned == ".":
        return None
    try:
        val = float(cleaned)
        if tok.isdigit() and 2 <= len(tok) <= 3:
            if expect_negative:
                val = -abs(val / 10.0)
            else:
                val = val / 10.0
        if expect_negative and not has_explicit_minus and not tok.startswith("-"):
            if tok not in _OCR_SIGN_MAP:
                val = -abs(val)
        return val
    except ValueError:
        return None


def _fix_bmd_token(tok: str) -> float:
    tok = tok.strip("[]|\"'(),")
    if re.match(r"^[01]\d{3}$", tok):
        return float(f"{tok[0]}.{tok[1:]}")
    return float(tok)


def _split_merged_ya_t(tok: str) -> tuple[Optional[int], Optional[float]]:
    m = _RE_MERGED_YA_T.match(tok.strip("[]|\"'(),"))
    if m:
        return int(m.group(1)), float(m.group(2)) / 10.0
    return None, None


def extract_bmd_row(row_str: str) -> BmdResult:
    tokens = row_str.split()
    bmd_idx = next(
        (i for i, tok in enumerate(tokens)
         if _RE_BMD_NEW.match(tok.strip("[]|\"'(),")) or _RE_BMD_OLD.match(tok.strip("[]|\"'(),"))),
        -1,
    )
    if bmd_idx == -1:
        return BmdResult()
    try:
        bmd = _fix_bmd_token(tokens[bmd_idx])
    except ValueError:
        return BmdResult()

    def _tok(offset: int) -> Optional[str]:
        idx = bmd_idx + offset
        return tokens[idx] if idx < len(tokens) else None

    t = z = None
    ya_tok = _tok(1)
    if ya_tok:
        merged_ya, merged_t = _split_merged_ya_t(ya_tok)
        if merged_ya is not None:
            t = merged_t
            z = parse_score_token(_tok(3)) if _tok(3) else None
        else:
            t = parse_score_token(_tok(2), expect_negative=True) if _tok(2) else None
            z = parse_score_token(_tok(4)) if _tok(4) else None

    def _clamp(v: Optional[float]) -> Optional[float]:
        return v if v is not None and SCORE_MIN <= v <= SCORE_MAX else None

    return BmdResult(bmd=bmd, t=_clamp(t), z=_clamp(z))


# ── GE Lunar section parser ───────────────────────────────────────────────────

def find_ancillary_section(txt: str, keyword: str) -> list[str]:
    kw_lower  = keyword.lower()
    txt_lower = txt.lower()

    start = next(
        (m.start() for m in _RE_ANCILLARY.finditer(txt_lower)
         if kw_lower in txt_lower[m.start(): m.start() + 60]),
        -1,
    )
    if start == -1:
        return []

    lines = txt[start: start + 3000].split("\n")
    header_idx = next(
        (i for i, line in enumerate(lines)
         if "region" in line.lower()
         and any(k in line.lower() for k in ("(g/cm", "(g/em", "(lem", "(olem", "(giem", "t-score"))),
        -1,
    )
    if header_idx == -1:
        return []

    data_rows: list[str] = []
    for line in lines[header_idx + 1:]:
        stripped = line.strip()
        if not stripped:
            continue
        if any(kw in stripped.lower() for kw in _SECTION_STOP_KEYWORDS):
            break
        if _digit_count(stripped) >= 2:
            data_rows.append(stripped)
        if len(data_rows) >= 50:
            break
    return data_rows


def ge_spine_l1l4(texts: dict[str, str]) -> BmdResult:
    for txt in texts.values():
        rows = find_ancillary_section(txt, "ap spine")
        if len(rows) >= 7:
            result = extract_bmd_row(rows[6])
            if result.is_valid():
                return result

    comb = combined(texts)
    for pattern in (
        r"(?:l1.?l4|spine).*?(\d{1,2}/\d{1,2}/(?:19|20)\d{2})\s+[0-9.]+\s+(0\.[4-9]\d{2}|1\.\d{3})",
        r"(\d{1,2}/\d{1,2}/(?:19|20)\d{2})\s+[0-9.]+\s+(0\.[4-9]\d{2}|1\.\d{3})\s+",
    ):
        m = re.search(pattern, comb, re.IGNORECASE)
        if m:
            bmd = float(m.group(2))
            if BMD_MIN < bmd < BMD_MAX:
                return BmdResult(bmd=bmd)
    return BmdResult()


def ge_femur(texts: dict[str, str], side: str) -> FemurResult:
    side        = side.lower()
    other_side  = "right" if side == "left" else "left"
    our_aliases = _SIDE_ALIASES[side]
    oth_aliases = _SIDE_ALIASES[other_side]

    result = FemurResult()
    for kw in (f"{side} femur", "dualfemur", "dual femur", "femur"):
        for txt in texts.values():
            rows = find_ancillary_section(txt, kw)
            if not rows:
                continue
            for row in rows:
                row_l   = row.lower()
                is_dual = kw.lower() in ("dualfemur", "dual femur", "femur")
                if is_dual:
                    has_ours  = any(a in row_l for a in our_aliases)
                    has_other = any(a in row_l for a in oth_aliases)
                    if (has_other and not has_ours) or "mean" in row_l or "dif" in row_l:
                        continue
                    if not has_ours:
                        continue
                r = extract_bmd_row(row)
                if not (r.bmd and BMD_MIN < r.bmd < BMD_MAX):
                    continue
                if "total" in row_l and result.total.bmd is None:
                    result.total = r
                elif "neck" in row_l and not any(w in row_l for w in ("upper", "lower", "ward")) and result.neck.bmd is None:
                    result.neck = r
        if result.neck.bmd or result.total.bmd:
            break
    return result


# ── Hologic parser ────────────────────────────────────────────────────────────

def parse_hologic(texts: dict[str, str]) -> dict[str, Optional[float]]:
    """
    Generic Hologic Horizon parser.
    Searches for standard region labels (L1-L4, Neck, Total) followed by
    BMD, T-score, and optionally Z-score on each line.
    """
    comb = combined(texts)

    spine = BmdResult()
    left  = FemurResult()
    right = FemurResult()

    for m in _RE_HOLOGIC_BMD_ROW.finditer(comb):
        label = m.group(0).split()[0].lower().strip(":-")
        try:
            bmd = float(m.group(1))
            t   = float(m.group(2)) if m.group(2) else None
            z   = float(m.group(3)) if m.group(3) else None
        except (ValueError, TypeError):
            continue

        if not (BMD_MIN < bmd < BMD_MAX):
            continue
        t = t if t is not None and SCORE_MIN <= t <= SCORE_MAX else None
        z = z if z is not None and SCORE_MIN <= z <= SCORE_MAX else None
        r = BmdResult(bmd=bmd, t=t, z=z)

        # Route by label
        context = comb[max(0, m.start() - 200): m.start()].lower()
        is_left  = "left"  in context or "left"  in label
        is_right = "right" in context or "right" in label

        if "l1" in label or "spine" in label:
            if not spine.bmd:
                spine = r
        elif "neck" in label:
            if is_left and not left.neck.bmd:
                left.neck = r
            elif is_right and not right.neck.bmd:
                right.neck = r
        elif "total" in label:
            if is_left and not left.total.bmd:
                left.total = r
            elif is_right and not right.total.bmd:
                right.total = r

    return {
        "Spine_L1L4_BMD":  spine.bmd,       "Spine_L1L4_T":  spine.t,       "Spine_L1L4_Z":  spine.z,
        "LFemur_Neck_BMD": left.neck.bmd,   "LFemur_Neck_T": left.neck.t,   "LFemur_Neck_Z": left.neck.z,
        "LFemur_Total_BMD":left.total.bmd,  "LFemur_Total_T":left.total.t,  "LFemur_Total_Z":left.total.z,
        "RFemur_Neck_BMD": right.neck.bmd,  "RFemur_Neck_T": right.neck.t,  "RFemur_Neck_Z": right.neck.z,
        "RFemur_Total_BMD":right.total.bmd, "RFemur_Total_T":right.total.t, "RFemur_Total_Z":right.total.z,
    }


# ── demographics parsers ──────────────────────────────────────────────────────

def _clean_str(s: Optional[str]) -> Optional[str]:
    if not s:
        return None
    s = _RE_WHITESPACE.sub(" ", s).strip()
    s = re.sub(r"^[^A-Za-z]+", "", s).strip()
    return s or None


def parse_age(txt: str) -> Optional[float]:
    m = _RE_AGE_LABEL.search(txt) or _RE_AGE_YEARS.search(txt)
    if not m:
        return None
    raw = m.group(1)
    val = float(raw)
    if val > 150:
        val = float(raw[:2] + "." + raw[2:])
    return round(val, 1)


def parse_dob(txt: str, study_year: Optional[int] = None, age: Optional[float] = None) -> Optional[str]:
    m = _RE_DOB.search(txt)
    if m:
        raw = m.group(1).strip()
        raw = re.sub(r"[oO](?=\d)|(?<=\d)[oO]", "0", raw)
        dm = _RE_DOB_DIGITS.search(raw)
        if dm:
            mo, day, yr = dm.groups()
            if len(yr) == 4 and int(yr) > 2025:
                yr = "19" + yr[2:]
            return f"{mo}/{day}/{yr}"
    if study_year and age:
        return f"1/1/{study_year - int(age)}"
    return None


def parse_name(txt: str) -> Optional[str]:
    m = _RE_NAME.search(txt)
    if not m:
        return None
    candidate = m.group(1)
    for stop in _NAME_STOPS:
        idx = candidate.lower().find(stop.lower())
        if idx != -1:
            candidate = candidate[:idx]
    cand = _clean_str(candidate)
    return None if (not cand or len(cand) < 4 or cand.upper() == cand) else cand


def parse_sex(txt: str) -> Optional[str]:
    m = _RE_SEX_LABEL.search(txt) or _RE_SEX_WORD.search(txt)
    return m.group(1)[0].upper() if m else None


def parse_height(txt: str) -> Optional[float]:
    m = _RE_HEIGHT_IN.search(txt)
    if m:
        raw = (m.group(1).replace(" ", "")
               .replace("s", "5").replace("S", "5")
               .replace("l", "1").replace("L", "1"))
        raw = re.sub(r"[oO](?=\d)|(?<=\d)[oO]", "0", raw)
        raw = re.sub(r"[^0-9.]", "", raw)
        if raw:
            val = float(raw)
            if val > 100:
                val /= 10.0
            if 48.0 <= val <= 84.0:
                return round(val, 1)
    m2 = _RE_HEIGHT_UNIT.search(txt)
    if m2:
        val = float(m2.group(1))
        return round(val / 10.0 if val > 100 else val, 1)
    return None


def parse_weight(txt: str) -> Optional[float]:
    m = _RE_WEIGHT.search(txt)
    if not m:
        return None
    val = float(m.group(1))
    return round(val / 10.0 if val > 500 else val, 1)


def parse_scan_date(txt: str) -> Optional[str]:
    m = _RE_SCAN_DATE.search(txt)
    if m:
        dm = _RE_DATE_MDY.search(m.group(1).strip().rstrip(","))
        return dm.group(1) if dm else m.group(1).strip()[:20]
    return None


def parse_physician(txt: str) -> Optional[str]:
    m = _RE_PHYSICIAN.search(txt)
    if not m:
        return None
    cand = m.group(1)
    for stop in ("Birth", "Date", "Height", "Weight", "Patient", "\n", "Facility"):
        idx = cand.lower().find(stop.lower())
        if idx != -1:
            cand = cand[:idx]
    return _clean_str(cand)


def parse_tbs(txt: str) -> Optional[float]:
    m = _RE_TBS.search(txt)
    return float(m.group(1)) if m else None


# ── per-patient orchestrator ──────────────────────────────────────────────────

def process_patient(folder_num: str) -> Optional[dict]:
    patient_dir = TEXT_DIR / f"Patient_{folder_num}"
    if not patient_dir.is_dir():
        return None

    texts = read_folder(patient_dir)
    if not texts:
        return None

    comb = combined(texts)

    rec: dict = {
        "Folder":       folder_num,
        "Name":         parse_name(comb),
        "DOB":          parse_dob(comb),
        "Age_years":    parse_age(comb),
        "Sex":          parse_sex(comb),
        "Height_in":    parse_height(comb),
        "Weight_lbs":   parse_weight(comb),
        "Referring_MD": parse_physician(comb),
        "Scan_Date":    parse_scan_date(comb),
        "TBS_L1L4":     parse_tbs(comb),
    }

    if "hologic" in comb.lower():
        bd = parse_hologic(texts)
    else:
        spine = ge_spine_l1l4(texts)
        left  = ge_femur(texts, "left")
        right = ge_femur(texts, "right")
        bd = {
            "Spine_L1L4_BMD":  spine.bmd,       "Spine_L1L4_T":  spine.t,       "Spine_L1L4_Z":  spine.z,
            "LFemur_Neck_BMD": left.neck.bmd,   "LFemur_Neck_T": left.neck.t,   "LFemur_Neck_Z": left.neck.z,
            "LFemur_Total_BMD":left.total.bmd,  "LFemur_Total_T":left.total.t,  "LFemur_Total_Z":left.total.z,
            "RFemur_Neck_BMD": right.neck.bmd,  "RFemur_Neck_T": right.neck.t,  "RFemur_Neck_Z": right.neck.z,
            "RFemur_Total_BMD":right.total.bmd, "RFemur_Total_T":right.total.t, "RFemur_Total_Z":right.total.z,
        }

    rec.update(bd)
    return rec


# ── Excel styling ─────────────────────────────────────────────────────────────

_THIN_SIDE    = Side(style="thin", color="CCCCCC")
_CELL_BORDER  = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
_EVEN_FILL    = PatternFill("solid", fgColor="EEF2FF")
_ODD_FILL     = PatternFill("solid", fgColor="FFFFFF")
_RED_FILL     = PatternFill("solid", fgColor="FFD7D7")
_YELLOW_FILL  = PatternFill("solid", fgColor="FFF3CD")

_CENTER = Alignment(horizontal="center")
_CENTER_WRAP = Alignment(horizontal="center", wrap_text=True)


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _CENTER_WRAP
        cell.border    = _CELL_BORDER


def _style_body(ws) -> None:
    t_col_indices = {
        cell.column for cell in ws[1] if cell.value and "_T" in str(cell.value)
    }
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
        base_fill = _EVEN_FILL if row_idx % 2 == 0 else _ODD_FILL
        for cell in row:
            cell.fill      = base_fill
            cell.alignment = _CENTER
            cell.border    = _CELL_BORDER
            if cell.column in t_col_indices and cell.value is not None:
                try:
                    v = float(cell.value)
                    if v <= T_OSTEOPOROSIS:
                        cell.fill = _RED_FILL
                    elif v <= T_OSTEOPENIA:
                        cell.fill = _YELLOW_FILL
                except (ValueError, TypeError):
                    pass


def _autofit_columns(ws) -> None:
    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 30)


# ── main ──────────────────────────────────────────────────────────────────────

_COL_ORDER = [
    "Folder", "Name", "DOB", "Age_years", "Sex",
    "Height_in", "Weight_lbs", "Referring_MD", "Scan_Date",
    "Spine_L1L4_BMD", "Spine_L1L4_T", "Spine_L1L4_Z",
    "LFemur_Neck_BMD", "LFemur_Neck_T", "LFemur_Neck_Z",
    "LFemur_Total_BMD", "LFemur_Total_T", "LFemur_Total_Z",
    "RFemur_Neck_BMD", "RFemur_Neck_T", "RFemur_Neck_Z",
    "RFemur_Total_BMD", "RFemur_Total_T", "RFemur_Total_Z",
    "TBS_L1L4",
]


def main() -> None:
    folders = sorted(
        (d.name.removeprefix("Patient_") for d in TEXT_DIR.iterdir() if d.name.startswith("Patient_")),
        key=lambda v: int(v) if v.isdigit() else 999,
    )

    rows = [rec for fn in folders if (rec := process_patient(fn)) is not None]
    df   = pd.DataFrame(rows)

    ordered = [c for c in _COL_ORDER if c in df.columns]
    ordered += [c for c in df.columns if c not in ordered]
    df = df[ordered]

    bmd_cols   = [c for c in df.columns if "BMD" in c]
    score_cols = [c for c in df.columns if "_T" in c or "_Z" in c or "TBS" in c]
    df[bmd_cols]   = df[bmd_cols].apply(pd.to_numeric, errors="coerce").round(3)
    df[score_cols] = df[score_cols].apply(pd.to_numeric, errors="coerce").round(1)

    with pd.ExcelWriter(OUTPUT_XLS, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="DXA Data", index=False)
        ws = writer.sheets["DXA Data"]
        _style_header(ws)
        _style_body(ws)
        _autofit_columns(ws)
        ws.freeze_panes = "A2"

    print(f"\n✅  Saved: {OUTPUT_XLS}")
    print(df.to_string(index=False))


if __name__ == "__main__":
    main()